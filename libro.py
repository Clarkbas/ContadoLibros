from openpyxl import Workbook

# Datos de los libros
books_data = [
    {"titulo": "To Kill a Mockingbird", "genero": "Ficción", "paginas": 281, "saga": "No"},
    {"titulo": "1984", "genero": "Ciencia ficción", "paginas": 328, "saga": "No"},
    {"titulo": "Pride and Prejudice", "genero": "Romance", "paginas": 279, "saga": "No"},
    {"titulo": "The Catcher in the Rye", "genero": "Ficción contemporánea", "paginas": 277, "saga": "No"},
    {"titulo": "The Great Gatsby", "genero": "Ficción literaria", "paginas": 180, "saga": "No"},
    {"titulo": "The Hobbit", "genero": "Fantasía", "paginas": 310, "saga": "Sí"},
    {"titulo": "The Hunger Games", "genero": "Ciencia ficción", "paginas": 374, "saga": "Sí"},
    {"titulo": "The Da Vinci Code", "genero": "Misterio", "paginas": 489, "saga": "Sí"},
    {"titulo": "The Shining", "genero": "Horror", "paginas": 447, "saga": "No"},
    {"titulo": "Gone with the Wind", "genero": "Histórico", "paginas": 1037, "saga": "No"},
    {"titulo": "The Fault in Our Stars", "genero": "Romance", "paginas": 313, "saga": "No"},
    {"titulo": "Dune", "genero": "Ciencia ficción", "paginas": 412, "saga": "Sí"},
    {"titulo": "Harry Potter and the Sorcerer's Stone", "genero": "Fantasía", "paginas": 309, "saga": "Sí"},
    {"titulo": "The Alchemist", "genero": "Ficción", "paginas": 208, "saga": "No"},
    {"titulo": "The Catcher in the Rye", "genero": "Ficción contemporánea", "paginas": 277, "saga": "No"},
    {"titulo": "Brave New World", "genero": "Ciencia ficción", "paginas": 311, "saga": "No"},
    {"titulo": "Moby-Dick", "genero": "Aventura", "paginas": 635, "saga": "No"},
    {"titulo": "War and Peace", "genero": "Histórico", "paginas": 1225, "saga": "No"},
    {"titulo": "The Odyssey", "genero": "Épico", "paginas": 500, "saga": "No"},
    {"titulo": "Les Misérables", "genero": "Histórico", "paginas": 1463, "saga": "No"},
    {"titulo": "The Chronicles of Narnia: The Lion, the Witch and the Wardrobe", "genero": "Fantasía", "paginas": 206, "saga": "Sí"},
    {"titulo": "Percy Jackson & the Olympians: The Lightning Thief", "genero": "Fantasía", "paginas": 377, "saga": "Sí"},
    {"titulo": "The Maze Runner", "genero": "Ciencia ficción", "paginas": 374, "saga": "Sí"},
    {"titulo": "Twilight", "genero": "Romance", "paginas": 498, "saga": "Sí"},
    {"titulo": "The Girl with the Dragon Tattoo", "genero": "Misterio", "paginas": 465, "saga": "Sí"},
    {"titulo": "A Game of Thrones", "genero": "Fantasía", "paginas": 835, "saga": "Sí"},
    {"titulo": "The Book Thief", "genero": "Histórico", "paginas": 584, "saga": "No"},
    {"titulo": "Life of Pi", "genero": "Aventura", "paginas": 319, "saga": "No"},
    {"titulo": "The Handmaid's Tale", "genero": "Distopía", "paginas": 311, "saga": "No"},
    {"titulo": "The Road", "genero": "Post-apocalíptico", "paginas": 287, "saga": "No"},
]

# Crear un nuevo libro de trabajo
wb = Workbook()
sheet = wb.active
sheet.title = "Libros"

# Encabezados de las columnas
sheet.cell(row=1, column=1, value="Título")
sheet.cell(row=1, column=2, value="Género")
sheet.cell(row=1, column=3, value="Páginas")
sheet.cell(row=1, column=4, value="Saga")

# Escribir los datos de los libros en el archivo Excel
for idx, libro in enumerate(books_data, start=2):
    sheet.cell(row=idx, column=1, value=libro["titulo"])
    sheet.cell(row=idx, column=2, value=libro["genero"])
    sheet.cell(row=idx, column=3, value=libro["paginas"])
    sheet.cell(row=idx, column=4, value="Sí" if libro["saga"] == "Sí" else "No")

# Contar la cantidad de sagas
cantidad_sagas = sum(1 for libro in books_data if libro["saga"] == "Sí")


# Guardar el archivo Excel
file_path = "C:/Cosas de Bastan/proyecto python/Libros.xlsx"
wb.save(file_path)
print(f"Archivo guardado en: {file_path}")
