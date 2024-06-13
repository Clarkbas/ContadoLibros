import csv
from openpyxl import load_workbook
import charts


def readExcel(path):
    wb = load_workbook(path)  # Carga el libro de Excel desde la ruta especificada.
    sheet = wb.active
    header = [cell.value for cell in sheet[1]] # Extrae la primera fila como encabezado.

    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_dict = {header[i]: row[i] for i in range(len(header))} # Crea un diccionario para cada fila, asociando los valores a los nombres del encabezado.
        data.append(row_dict)

    return data


def countSagas(data):
    return sum(1 for libro in data if libro["Saga"] == "Sí") # Cuenta la cantidad de libros en los que el valor de la clave "Saga" es "Sí"


def classifyBooks(data):
    genres = set([libro["Género"] for libro in data]) # Crea un conjunto de géneros únicos a partir de los datos.
    genre_counts = {genre: sum(1 for libro in data if libro["Género"] == genre) for genre in genres}
    return genre_counts # Luego los retorna.


if __name__ == '__main__':
    data = readExcel('C:/Cosas de Bastan/proyecto python/Libros.xlsx')

    print("Cantidad de sagas:", countSagas(data)) # Imprime la cantidad de libros que son parte de sagas.

    genre_counts = classifyBooks(data)
    print("Cantidad de libros por género:")
    for genre, count in genre_counts.items():  # Itera sobre el diccionario y imprime la cantidad de libros por género.
        print(f"{genre}: {count}")

    labels = list(genre_counts.keys())
    values = list(genre_counts.values())
    charts.generateBarChart(labels, values) # Genera un gráfico de barras utilizando las etiquetas y valores obtenidos.
